import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.marker import DataPoint
import random


def generate_slices_colors(num_slices):
    # Generate random colors for slices
    colors = ['#' + ''.join(random.choices('0123456789ABCDEF', k=6)) for _ in range(num_slices)]
    slices = [DataPoint(idx=i) for i in range(num_slices)]
    for i, slice_point in enumerate(slices):
        slice_point.graphicalProperties.solidFill = colors[i]
    return slices

def process_workbook(filename):
    wb = xl.load_workbook(filename)

    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    corrected_price_cell_heading = sheet.cell(1, 4)

    corrected_price_cell_heading.value = 'corrected price'

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    # Use the reusable function to generate slices with random colors

    slices = generate_slices_colors(3)

    chart.series[0].data_points = slices



    wb.save(filename)



process_workbook('transaction3.xlsx')

import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.marker import DataPoint
import random


def generate_slices_colors(num_slices):
    # Generate random colors for slices
    colors = ['#' + ''.join(random.choices('0123456789ABCDEF', k=6)) for _ in range(num_slices)]
    slices = [DataPoint(idx=i) for i in range(num_slices)]
    for i, slice_point in enumerate(slices):
        slice_point.graphicalProperties.solidFill = colors[i]
    return slices

def process_workbook(filename):
    wb = xl.load_workbook(filename)

    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    corrected_price_cell_heading = sheet.cell(1, 4)

    corrected_price_cell_heading.value = 'corrected price'

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    # Use the reusable function to generate slices with random colors

    slices = generate_slices_colors(3)

    chart.series[0].data_points = slices



    wb.save(filename)



process_workbook('transaction3.xlsx')

